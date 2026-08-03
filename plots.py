"""Reusable panel-drawing functions for the Joule composite figures."""
import csv, json, warnings
warnings.filterwarnings("ignore")
import numpy as np
from ccus_style import plt, panel, despine, INK, MUT, BLUE, TEAL, ORANGE, VERM, GREY, GRID
from matplotlib.patches import Rectangle, FancyBboxPatch, FancyArrowPatch
from matplotlib.lines import Line2D

D = "/home/claude/ccus_uhs"

def read_psv(path):
    with open(path, encoding="utf-8") as f:
        rows = [r for r in csv.reader(f, delimiter="|") if r and any(x.strip() for x in r)]
    return rows[0], rows[1:]

# ---------------- shared data ----------------
def load_all():
    _, gc = read_psv(f"{D}/data/geo_ccus.psv")
    geo_ccus = {r[0]: (float(r[2]), float(r[3]), r[4]) for r in gc if r[2] != "NA"}
    with open(f"{D}/data/geo_uhs_blueh2.psv") as f:
        txt = f.read()
    def pb(block):
        lines = [l for l in block.strip().splitlines() if l.strip()]
        return {l.split("|")[0]: l.split("|")[1:6] for l in lines[1:]}
    geo_uhs = pb(txt.split("===UHS===")[1].split("===BLUEH2===")[0])
    geo_bh = pb(txt.split("===BLUEH2===")[1])
    ccus_attr = {}
    for fn in ["ccus_americas.psv", "ccus_emea.psv", "ccus_apac.psv"]:
        h, dd = read_psv(f"{D}/data/{fn}")
        for r in dd:
            ccus_attr[r[0]] = dict(zip(h, r))
    return geo_ccus, geo_uhs, geo_bh, ccus_attr

def active(st):
    s = st.lower()
    return not ("cancel" in s or "suspend" in s)

# ---------------- map panels ----------------
import geopandas as gpd
_world = gpd.read_file(gpd.datasets.get_path("naturalearth_lowres"))

def basemap(ax, xlim, ylim):
    _world.plot(ax=ax, color="#EDEAE3", edgecolor="white", linewidth=0.5, zorder=0)
    ax.set_xlim(xlim); ax.set_ylim(ylim)
    ax.set_xticks([]); ax.set_yticks([])
    for s in ax.spines.values():
        s.set_visible(True); s.set_color("#B8B4AC"); s.set_linewidth(0.6)

def map_layers(ax, geo_ccus, geo_uhs, geo_bh, ccus_attr, smin=6, smax=140, alpha=0.75, lw=0.5):
    pts = []
    for name, (lat, lon, basin) in geo_ccus.items():
        a = ccus_attr.get(name, {})
        try: c = float(a.get("capture_capacity_MtCO2_per_yr", "NA"))
        except ValueError: c = 0.3
        pts.append((lon, lat, c, active(a.get("status", ""))))
    inact = [(x, y) for x, y, c, a in pts if not a]
    act = [(x, y, c) for x, y, c, a in pts if a]
    ax.scatter([p[0] for p in inact], [p[1] for p in inact], s=smin, c="#C4C0B8",
               marker="o", alpha=0.6, linewidths=0, zorder=2)
    ax.scatter([p[0] for p in act], [p[1] for p in act],
               s=[min(smax, smin + p[2] * 9) for p in act],
               c=BLUE, marker="o", alpha=alpha, edgecolor="white", linewidths=lw, zorder=3)
    bh = [(float(v[2]), float(v[1])) for v in geo_bh.values()]
    uh = [(float(v[2]), float(v[1])) for v in geo_uhs.values()]
    ax.scatter([p[0] for p in bh], [p[1] for p in bh], s=int(smax*0.24), c=ORANGE,
               marker="D", alpha=0.9, edgecolor="white", linewidths=lw, zorder=4)
    ax.scatter([p[0] for p in uh], [p[1] for p in uh], s=int(smax*0.30), c=TEAL,
               marker="^", alpha=0.95, edgecolor="white", linewidths=lw, zorder=5)

ZOOMS = [
    ("US Gulf Coast", (-98.5, -87.5), (26.0, 33.5)),
    ("Alberta–Williston", (-121, -99), (47.5, 57.5)),
    ("Northwest Europe", (-6.5, 13), (49.5, 62.5)),
    ("Eastern China", (104, 123.5), (28.5, 42)),
]
ZOOM_NOTES = {
    0: [("Spindletop · Moss Bluff · Clemens\nH2 caverns amid CO2 storage hubs", -93.0, 26.6, "center")],
    1: [("Industrial Heartland: Quest ·\nACTL · H2 complex + Lotsberg salt", -120.3, 48.2, "left"),
        ("Weyburn–Midale /\nBoundary Dam", -99.8, 51.8, "right")],
    2: [("North Sea stores +\nZechstein salt caverns", -6.2, 60.0, "left"),
        ("HyNet ·\nCheshire", -6.3, 49.9, "left")],
    3: [("Bohai Bay EOR/CCS +\nJintan & Pingdingshan salt", 105.0, 29.0, "left")],
}

def map_legend_handles():
    return [
        Line2D([], [], marker="o", ls="", mfc=BLUE, mec="white", ms=7, label="CCUS project (size = Mt CO$_2$/yr; active pipeline)"),
        Line2D([], [], marker="o", ls="", mfc="#C4C0B8", mec="none", ms=4.5, label="CCUS cancelled/suspended"),
        Line2D([], [], marker="^", ls="", mfc=TEAL, mec="white", ms=7, label="Underground H$_2$ storage"),
        Line2D([], [], marker="D", ls="", mfc=ORANGE, mec="white", ms=6, label="H$_2$/NH$_3$ production with CCS"),
    ]

# ---------------- chart panels ----------------
def draw_region_status(ax, ccus_attr):
    def sg(s):
        s = (s or "").lower()
        if "cancel" in s or "suspend" in s: return "Cancelled / suspended"
        if "pilot" in s or "injection completed" in s: return "Pilot / completed demo"
        if "operational" in s: return "Operational"
        if "construction" in s: return "Under construction"
        if "advanced" in s: return "Advanced development"
        return "Planned"
    regions_map = {"ccus_americas.psv": "Americas", "ccus_emea.psv": "Europe–ME–Africa", "ccus_apac.psv": "Asia–Pacific"}
    order = ["Operational", "Under construction", "Advanced development", "Planned", "Pilot / completed demo", "Cancelled / suspended"]
    counts = {r: {o: 0 for o in order} for r in regions_map.values()}
    for fn, rg in regions_map.items():
        h, dd = read_psv(f"{D}/data/{fn}")
        for r in dd:
            counts[rg][sg(r[3])] += 1
    scol = {"Operational": "#08519C", "Under construction": "#3182BD",
            "Advanced development": "#6BAED6", "Planned": "#C6DBEF",
            "Pilot / completed demo": TEAL, "Cancelled / suspended": "#CFCBC4"}
    regs = list(regions_map.values())
    y = np.arange(len(regs))[::-1]
    left = np.zeros(len(regs))
    for st in order:
        vals = np.array([counts[r][st] for r in regs], float)
        hatch = "///" if st == "Cancelled / suspended" else None
        ax.barh(y, vals, left=left, height=0.6, color=scol[st], hatch=hatch,
                edgecolor="white", linewidth=1.0, label=st)
        for yi, v, l in zip(y, vals, left):
            if v >= 8:
                tc = "white" if st in ("Operational", "Under construction") else INK
                ax.text(l + v / 2, yi, int(v), ha="center", va="center", fontsize=7.8, color=tc)
        left += vals
    ax.set_yticks(y); ax.set_yticklabels(regs, fontsize=8.5)
    ax.set_xlabel("CCUS projects (n = 242)")
    ax.set_xlim(0, 86)
    despine(ax, keep=("bottom",))
    ax.tick_params(axis="y", length=0)
    ax.legend(ncol=2, frameon=False, loc="upper center", bbox_to_anchor=(0.5, -0.30),
              handlelength=1.1, columnspacing=0.8, fontsize=7.3)

def draw_country_scatter(ax):
    from openpyxl import load_workbook
    wb = load_workbook(f"{D}/Global_CCUS_UHS_Coupling_Database_v3.xlsx", data_only=True)
    cc = wb["Country_Coupling"]
    ctab = [[cc.cell(r, c).value for c in range(1, 6)] for r in range(5, cc.max_row + 1)]
    full = [c for c in ctab if c[1] and c[3] and c[4]]
    both = [c for c in ctab if c[1] and c[3] and not c[4]]
    onlyc = [c for c in ctab if c[1] and not c[3]]
    onlyu = [c for c in ctab if not c[1] and c[3]]
    for grp, col, lab in [(onlyc, "#CFCBC4", "CCUS only"), (onlyu, ORANGE, "UHS only"),
                          (both, TEAL, "CCUS + UHS"), (full, BLUE, "Full chain")]:
        xs = [max(c[1], 0.55) for c in grp]; ys = [c[3] or 0 for c in grp]
        ss = [18 + 20 * (c[4] or 0) for c in grp]
        ax.scatter(xs, ys, s=ss, c=col, alpha=0.85, edgecolor="white", linewidth=0.7, label=lab, zorder=3)
    ann = {"USA": (54, 4), "China": (29, 2), "Canada": (21, 1), "United Kingdom": (15, 5),
           "Germany": (4, 9), "France": (5, 5), "Netherlands": (9, 1), "Australia": (16, 1)}
    short = {"United Kingdom": "UK", "Netherlands": "NL", "Australia": "AUS", "Germany": "DEU",
             "France": "FRA", "Canada": "CAN", "USA": "USA", "China": "CHN"}
    off = {"USA": (0, 5), "China": (0, 5), "Canada": (9, -15), "United Kingdom": (11, 4), "Germany": (8, 3),
           "France": (-16, 4), "Netherlands": (-6, 8), "Australia": (0, 9)}
    for k, (x_, y_) in ann.items():
        dx, dy = off[k]
        ax.annotate(short[k], (x_, y_), textcoords="offset points", xytext=(dx, dy),
                    fontsize=7.5, color=INK, ha="center")
    ax.set_xscale("log"); ax.set_xlim(0.45, 80); ax.set_ylim(-0.6, 10.2)
    ax.set_xticks([1, 3, 10, 30]); ax.set_xticklabels(["1", "3", "10", "30"])
    ax.set_xlabel("CCUS projects per country (log)")
    ax.set_ylabel("UHS projects per country")
    despine(ax)
    ax.grid(True, which="major", color=GRID, linewidth=0.6, zorder=0)
    leg = ax.legend(frameon=False, loc="upper right", handlelength=0.9, fontsize=7.3,
              borderaxespad=0.2, labelspacing=0.3,
              title="Bubble size = H$_2$+CCS projects", title_fontsize=7)
    leg.get_title().set_color(MUT)
    
def draw_ladder(ax):
    rows = [
        ("Grey SMR", 9.2, 11.0, 14.0, "#6E6E6E"),
        ("Blue H$_2$, as commonly built\n(55–90% capture, ~1–2% CH$_4$ leakage)", 5.0, 6.5, 9.3, "#7FABD3"),
        ("Blue H$_2$, best practice\n(>90% capture, ≤1% leakage)", 0.8, 3.0, 4.0, BLUE),
        ("Turquoise (pyrolysis)", 0.8, 6.0, 9.9, TEAL),
        ("Green electrolysis\n(dedicated renewables, incl. embodied)", 0.5, 2.9, 4.6, "#2A9D3E"),
    ]
    ypos = np.arange(len(rows))[::-1]
    for (lab, lo, mid, hi, col), yy in zip(rows, ypos):
        ax.plot([lo, hi], [yy, yy], lw=6, color=col, alpha=0.35, solid_capstyle="round", zorder=2)
        ax.plot([lo, hi], [yy, yy], lw=1.4, color=col, solid_capstyle="round", zorder=3)
        ax.scatter([mid], [yy], s=40, color=col, edgecolor="white", linewidth=1.0, zorder=4)
        ax.text(hi + 0.25, yy, f"{lo:g}–{hi:g}", va="center", fontsize=8.1, color=MUT)
    ax.scatter([16.4], [ypos[1]], marker="D", s=28, facecolor="none", edgecolor=VERM, linewidth=1.2, zorder=4)
    ax.annotate("GWP$_{20}$, 3.5% leakage view\n(Howarth & Jacobson 2021; contested)",
                (16.4, ypos[1]), textcoords="offset points", xytext=(-4, 10), fontsize=7.5,
                color=VERM, ha="center")
    for x_ in (3.38, 4.0):
        ax.axvline(x_, color="#B8B8B8", lw=0.8, ls=(0, (4, 3)), zorder=1)
    ax.text(3.30, len(rows) - 0.42, "EU 3.38", fontsize=7.5, color=MUT, ha="right")
    ax.text(4.08, len(rows) - 0.42, "US 45V 4.0", fontsize=7.5, color=MUT, ha="left")
    ax.annotate("", xy=(3.28, ypos[2] - 0.38), xytext=(3.0, ypos[2] - 0.38),
                arrowprops=dict(arrowstyle="-|>", color=BLUE, lw=1.0))
    ax.text(3.45, ypos[2] - 0.40, "+0.23–0.28 kg per kg H$_2$ cycled through UHS\n(storage adds only 2–10% to best-practice blue H$_2$)",
            fontsize=7.5, color=BLUE, va="center")
    ax.set_yticks(ypos); ax.set_yticklabels([r[0] for r in rows], fontsize=8.5)
    ax.set_xlabel("Life-cycle GWP (kg CO$_2$e per kg H$_2$, GWP$_{100}$ unless noted)")
    ax.set_xlim(0, 18.5); ax.set_ylim(-0.7, len(rows) - 0.2)
    despine(ax, keep=("bottom",))
    ax.tick_params(axis="y", length=0)
    ax.grid(axis="x", color=GRID, linewidth=0.6, zorder=0)

def draw_schematic(ax):
    """Quantified coupling framework: value chain with mass balance + archetype badges."""
    ax.set_xlim(0, 100); ax.set_ylim(0, 62); ax.axis("off")
    def box(x, y, w, h, lines, fc, ec, fs=8.3, bold_first=True, r=1.6, tc=INK):
        ax.add_patch(FancyBboxPatch((x, y), w, h, boxstyle=f"round,pad=0.7,rounding_size={r}",
                                    facecolor=fc, edgecolor=ec, linewidth=1.0, zorder=2))
        n = len(lines)
        for i, (txt, b) in enumerate(lines):
            yy = y + h * (n - i - 0.5) / n
            ax.text(x + w / 2, yy, txt, ha="center", va="center", fontsize=fs,
                    color=tc, zorder=3, fontweight="bold" if (b and bold_first) else "normal",
                    linespacing=1.15)
    def arrow(x1, y1, x2, y2, color, lw=1.5, style="-|>", ls="-", ms=10):
        ax.add_patch(FancyArrowPatch((x1, y1), (x2, y2), arrowstyle=style, mutation_scale=ms,
                                     color=color, linewidth=lw, linestyle=ls, zorder=4))
    def badge(x, y, num, color, label):
        ax.add_patch(plt.Circle((x, y), 1.9, facecolor=color, edgecolor="white", lw=1.0, zorder=6))
        ax.text(x, y, num, ha="center", va="center", fontsize=9.1, color="white",
                fontweight="bold", zorder=7)
        ax.text(x + 2.8, y, label, ha="left", va="center", fontsize=7.5, color=color,
                fontweight="bold", zorder=7)
    GREYE = "#8A8A8A"
    # ---- top row: the value chain ----
    TY, TH = 47, 11
    box(1.5, TY, 15, TH, [("Natural gas", True), ("upstream CH$_4$", False), ("leakage 0.2–8%", False)], "#F4F1EC", GREYE)
    box(21.5, TY, 17, TH, [("Reformer", True), ("ATR/SMR with", False), (">90% CO$_2$ capture", False)], "#EAF2FA", BLUE)
    box(43.5, TY, 15, TH, [("Low-carbon H$_2$", True), ("0.8–4 kg CO$_2$e", False), ("per kg H$_2$", False)], "#EAF2FA", BLUE)
    box(63.5, TY, 16, TH, [("Underground H$_2$", True), ("storage (UHS)", True), ("+0.23–0.28 kg CO$_2$e", False)], "#E7F5F0", TEAL)
    box(84.5, TY, 14, TH, [("End use", True), ("power, industry,", False), ("mobility", False)], "#F4F1EC", GREYE)
    arrow(17.2, TY + TH/2, 20.8, TY + TH/2, GREYE)
    arrow(39.2, TY + TH/2, 42.8, TY + TH/2, BLUE)
    # two-way seasonal cycling between H2 and UHS
    arrow(59.2, TY + TH/2 + 1.5, 62.8, TY + TH/2 + 1.5, TEAL)
    arrow(62.8, TY + TH/2 - 1.5, 59.2, TY + TH/2 - 1.5, TEAL)
    ax.text(61.0, TY + TH + 1.5, "seasonal\ncycling", fontsize=7, color=MUT, ha="center", va="bottom")
    arrow(80.2, TY + TH/2, 83.8, TY + TH/2, TEAL)
    # ---- CO2 branch ----
    arrow(30, TY - 0.6, 30, 33.8, BLUE, lw=2.0)
    ax.text(31.2, 40.5, "$\\approx$8–9 kg CO$_2$ captured", fontsize=7.5, color=BLUE, ha="left")
    ax.text(31.2, 38.2, "per kg H$_2$", fontsize=7.5, color=BLUE, ha="left")
    box(21.5, 22.5, 17, 10.5, [("Geological CO$_2$", True), ("storage", True), ("permanent", False)], "#EAF2FA", BLUE)
    # ---- archetype 3: cushion gas ----
    arrow(39.2, 27.5, 69, 45.8, ORANGE, lw=1.5, ls=(0, (5, 3)))
    ax.text(48.3, 33.6, "$\\approx$ +8% working H$_2$ capacity", fontsize=7.5, color="#B25000", ha="left")
    # ---- shared substrate band ----
    ax.add_patch(FancyBboxPatch((3, 2.5), 94, 13.5, boxstyle="round,pad=0.7,rounding_size=2.2",
                                facecolor="#FBF6EC", edgecolor="#D9C89A", linewidth=1.0, zorder=1))
    ax.text(50, 13.3, "SHARED BASIN RESOURCES", fontsize=9.1,
            color="#7A6420", ha="center", fontweight="bold", zorder=3)
    shared = ["Pore space and\ncaprock integrity", "Wells and drilling\ncapacity", "Pipelines and\nrights-of-way",
              "Monitoring and\nMMV networks", "Subsurface data", "Regulation and\nsocial licence"]
    for i, t in enumerate(shared):
        ax.text(11.5 + i * 15.4, 7.0, t, fontsize=7, color="#5C4D1A", ha="center", va="center", zorder=3)
    arrow(30, 21.7, 30, 17.0, "#C4A94E", lw=1.1, style="-")
    arrow(71.5, TY - 0.6, 71.5, 17.0, "#C4A94E", lw=1.1, style="-")
    # ---- archetype badges ----
    badge(6.5, 19.5, "1", "#7A6420", "Shared-basin co-location")
    badge(6.5, 59.5, "2", BLUE, "Integrated value chain: CO$_2$ down, H$_2$ cycling")
    badge(45.5, 36.5, "3", "#B25000", "CO$_2$ cushion gas")

A = json.load(open(f"{D}/analysis.json"))
K = json.load(open(f"{D}/km.json"))

def draw_scurves(ax1, ax2):
    years = np.array(A["years"])
    ax1.fill_between(years, A["ccus_cum_cap"], color=BLUE, alpha=0.18, lw=0)
    ax1.plot(years, A["ccus_cum_cap"], color=BLUE, lw=1.6)
    ax1.set_ylabel("CCUS capture\n(Mt CO$_2$/yr)", fontsize=8.1)
    ax1.text(1994.6, 32, "Sleipner (1996)", fontsize=7.3, color=MUT, ha="right")
    ax1.axvline(1996, color="#CCCCCC", lw=0.7, ls=":")
    ax1.tick_params(labelbottom=False)
    ax2.fill_between(years, A["uhs_cum"], color=TEAL, alpha=0.18, lw=0)
    ax2.plot(years, A["uhs_cum"], color=TEAL, lw=1.6)
    ax2.set_ylabel("UHS sites\n(cumulative)", fontsize=8.1)
    ax2.axvline(2020, color="#CCCCCC", lw=0.7, ls=":")
    ax2.text(2018.6, 14, "pure-H$_2$ wave\n(2020– )", fontsize=7.3, color=MUT, ha="right")
    for ax in (ax1, ax2):
        despine(ax)
        ax.grid(axis="y", color=GRID, lw=0.5)
        ax.set_xlim(1970, 2026)
    ax2.set_xlabel("Year")

def draw_cdf(ax):
    def cdf(ds, color, label):
        x = np.sort(ds); y = 100 * np.arange(1, len(x) + 1) / len(x)
        x = np.clip(x, 1, None)
        ax.step(x, y, where="post", color=color, lw=1.6, label=label)
    cdf(A["uhs_dists"], TEAL, "UHS site → nearest active CCUS")
    cdf(A["bh_dists"], ORANGE, "H$_2$+CCS plant → nearest UHS")
    ax.set_xscale("log"); ax.set_xlim(1, 4000); ax.set_ylim(0, 100)
    ax.axvline(100, color="#CCCCCC", lw=0.7, ls=":")
    ax.text(105, 5, "100 km", fontsize=7.3, color=MUT)
    ax.scatter([139], [50], s=18, color=TEAL, zorder=5, edgecolor="white", lw=0.7)
    ax.annotate("median 139 km", (139, 50), textcoords="offset points", xytext=(9, -15), fontsize=7.5, color="#00795A")
    ax.scatter([296], [50], s=18, color=ORANGE, zorder=5, edgecolor="white", lw=0.7)
    ax.annotate("median 296 km", (296, 50), textcoords="offset points", xytext=(5, 4), fontsize=7.5, color="#B25000")
    ax.set_xlabel("Distance (km, log)"); ax.set_ylabel("Share of sites (%)")
    despine(ax)
    ax.grid(color=GRID, lw=0.5)
    ax.legend(frameon=False, loc="upper left", bbox_to_anchor=(0.0, 1.04), handlelength=1.2, fontsize=7.3)

def draw_attrition(ax):
    attr = A["attrition"]
    order = sorted(attr.items(), key=lambda kv: kv[1][2])
    labels = [k for k, _ in order]; pct = [v[2] for _, v in order]
    ns = [f"{v[0]}/{v[1]}" for _, v in order]
    cols = [VERM if l == "Power" else "#9DBFD8" for l in labels]
    y = np.arange(len(labels))
    ax.barh(y, pct, color=cols, height=0.6, edgecolor="white", lw=0.7)
    for yi, p, n in zip(y, pct, ns):
        ax.text(p + 0.8, yi, f"{p}%  ({n})", va="center", fontsize=7.3, color=INK)
    ax.set_yticks(y); ax.set_yticklabels(labels, fontsize=7.5)
    ax.set_xlabel("Cancelled or suspended (% of projects)")
    ax.set_xlim(0, 50)
    despine(ax, keep=("bottom",))
    ax.tick_params(axis="y", length=0)
    ax.grid(axis="x", color=GRID, lw=0.5)

def draw_basins(ax):
    basins = ["US Gulf Coast", "North Sea", "Alberta (WCSB)", "NW Australia", "Ordos", "Williston", "Arabian Platform", "East Irish Sea"]
    ccus_n = [17, 15, 14, 9, 8, 7, 7, 5]
    uhs_n = [3, 1, 1, 0, 0, 0, 0, 0]
    bh_n = [6, 5, 4, 0, 0, 1, 3, 1]
    y = np.arange(len(basins))[::-1]; h = 0.26
    ax.barh(y + h, ccus_n, height=h, color=BLUE, label="Active CCUS", edgecolor="white", lw=0.5)
    ax.barh(y, [u if u else 0.12 for u in uhs_n], height=h, color=TEAL, label="UHS (in basin)", edgecolor="white", lw=0.5)
    ax.barh(y - h, [b if b else 0.12 for b in bh_n], height=h, color=ORANGE, label="H$_2$+CCS", edgecolor="white", lw=0.5)
    for yi, (c, u, b) in zip(y, zip(ccus_n, uhs_n, bh_n)):
        ax.text(c + 0.3, yi + h, c, va="center", fontsize=7, color=MUT)
        ax.text((u if u else 0.12) + 0.3, yi, u, va="center", fontsize=7, color=MUT)
        ax.text((b if b else 0.12) + 0.3, yi - h, b, va="center", fontsize=7, color=MUT)
    ax.set_yticks(y); ax.set_yticklabels(basins, fontsize=7.5)
    ax.set_xlabel("Projects in this database (n)")
    ax.set_xlim(0, 20)
    despine(ax, keep=("bottom",))
    ax.tick_params(axis="y", length=0)
    ax.grid(axis="x", color=GRID, lw=0.5)
    ax.legend(frameon=False, loc="lower right", handlelength=1.0, fontsize=7)

def draw_km(ax):
    km_t, km_s = K["km_t"], K["km_s"]
    # recompute CI from survival data for the band
    import csv as _csv
    rows = []
    with open(f"{D}/data/survival.psv") as f:
        rr = [x for x in _csv.reader(f, delimiter="|") if x and any(c.strip() for c in x)]
    h = rr[0]
    for x in rr[1:]:
        rows.append(dict(zip(h, x)))
    power = [x for x in rows if x["group"] == "power"]
    data = sorted((max(int(x["event_year"]) - int(x["announcement_year"]), 0.5),
                   1 if x["db_outcome"] == "death" else 0) for x in power)
    times = sorted({t for t, d in data if d == 1})
    S, V = 1.0, 0.0
    kt, ks, klo, khi = [0.0], [1.0], [1.0], [1.0]
    for t in times:
        n_risk = sum(1 for tt, _ in data if tt >= t)
        d_t = sum(1 for tt, dd in data if tt == t and dd == 1)
        S *= (1 - d_t / n_risk)
        if n_risk > d_t: V += d_t / (n_risk * (n_risk - d_t))
        se = S * np.sqrt(V)
        kt.append(t); ks.append(S)
        klo.append(max(0, S - 1.96 * se)); khi.append(min(1, S + 1.96 * se))
    def stepify(t, s):
        xs, ys = [t[0]], [s[0]]
        for i in range(1, len(t)):
            xs += [t[i], t[i]]; ys += [s[i-1], s[i]]
        return xs, ys
    xs, ys = stepify(kt, ks); xlo, ylo = stepify(kt, klo); xhi, yhi = stepify(kt, khi)
    ax.fill_between(xlo, ylo, yhi, color=BLUE, alpha=0.14, lw=0)
    ax.plot(xs, ys, color=BLUE, lw=1.5)
    for t, d in data:
        if d == 0:
            s_at = next(ks[i] for i in range(len(kt)-1, -1, -1) if kt[i] <= t)
            ax.plot([t], [s_at], marker="|", color=BLUE, ms=4, mew=0.9, alpha=0.6)
    ax.set_xlim(0, 21); ax.set_ylim(0, 1.02)
    ax.set_xlabel("Years since project announcement")
    ax.set_ylabel("Survival probability")
    i5 = max(i for i in range(len(kt)) if kt[i] <= 5)
    i10 = max(i for i in range(len(kt)) if kt[i] <= 10)
    ax.annotate(f"S(5 yr) = {ks[i5]:.2f}", (5, ks[i5]), textcoords="offset points",
                xytext=(7, 7), fontsize=7.3, color=INK,
                arrowprops=dict(arrowstyle="-", color="#AAAAAA", lw=0.5))
    ax.annotate(f"S(10 yr) = {ks[i10]:.2f}", (10, ks[i10]), textcoords="offset points",
                xytext=(7, 7), fontsize=7.3, color=INK,
                arrowprops=dict(arrowstyle="-", color="#AAAAAA", lw=0.5))
    ax.text(0.98, 0.05, "Power-sector CCS cohort\nn = 56, events = 22\nshaded: 95% CI (Greenwood)\nticks: censored",
            transform=ax.transAxes, fontsize=6.8, color=MUT, ha="right", va="bottom")
    despine(ax)
    ax.grid(axis="y", color=GRID, lw=0.5)

def draw_strip(ax):
    groups = [("Power failures (n = 22)", K["power_deaths"], VERM, 1.0),
              ("Non-power failures (n = 14)", K["nonpower_deaths"], "#8A8A8A", 0.0)]
    for label, vals, col, base in groups:
        cnt = {}
        for v in sorted(vals):
            k = round(v * 2) / 2
            cnt[k] = cnt.get(k, 0) + 1
            off = (cnt[k] - 1) * 0.10 * (1 if cnt[k] % 2 == 0 else -1)
            ax.scatter([v], [base + off], s=20, color=col, alpha=0.85,
                       edgecolor="white", linewidths=0.5, zorder=3)
        med = float(np.median(vals))
        ax.plot([med, med], [base - 0.30, base + 0.30], color=INK, lw=1.1, zorder=4)
        ax.text(med, base + 0.38, f"median {med:g} yr", fontsize=7.3, color=INK, ha="center")
    ax.set_yticks([0, 1]); ax.set_yticklabels(["Non-power\n(n = 14)", "Power\n(n = 22)"], fontsize=7.5)
    ax.set_ylim(-0.7, 1.75); ax.set_xlim(0, 12.5)
    ax.set_xlabel("Years from announcement to termination")
    despine(ax, keep=("bottom",))
    ax.tick_params(axis="y", length=0)
    ax.grid(axis="x", color=GRID, lw=0.5)
